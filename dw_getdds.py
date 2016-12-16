#! /usr/bin/python
# Eric Gullufsen
# This script grabs Heating Degree Day for Juneau, AK for the period 09/01/2016 - 11/30/2016

# 'requests' is a nice library for making HTTP requests
import requests
import time
import openpyxl

API_KEY = 'cca5724c8d2d37e7'
URL_PREFIX = 'http://api.wunderground.com/api/'
URL_SUFFIX1 = '/history_'
URL_SUFFIX2 = '/q/AK/Juneau.json'

def get_degree_days():
	# construct list of strings that represent the days we want, in order
	         # List Comprehensions are rad! This one gives us
		        # ['01','02','03',...,'29','30',31']
	long_month_days = [('' if len(str(x)) == 2 else '0') + str(x) for x in range(1,32)]	
	short_month_days =  [('' if len(str(x)) == 2 else '0') + str(x) for x in range(1,31)]
	# we'll be lazy & just hard-code the month strings needed
	months = ['09','10','11']
	days_for_months = [short_month_days, long_month_days, short_month_days]

	year = '2016'
	cap_counter = 1
	spsh = openpyxl.Workbook()
	sheet = spsh.active
	
	for c1 in range(4):
		mo = months[c1]
		dfm = days_for_months[c1]
		for c2 in range(len(dfm)):
			date_string = year + mo + dfm[c2]
			get_url = URL_PREFIX + API_KEY + URL_SUFFIX1 + date_string + URL_SUFFIX2
			
			# if at our max of 10 requests per minute, delay execution for a minute
			if (cap_counter % 10 == 0):
				time.sleep(60)
			
			req = requests.get(get_url)
			jsonny = req.json()
			hdd = int(jsonny['history']['dailysummary'][0]['heatingdegreedays'])
			print "Date: " + mo + "/" + dfm[c2] + "/" + year + "\t" + "HDD: " + str(hdd)
			sheet.cell(row=cap_counter, column=1).value = mo + "/" + dfm[c2] + "/" + year
			sheet.cell(row=cap_counter, column=2).value = hdd
			# print get_url
			cap_counter += 1
	spsh.save('DDericg.xlsx')
if __name__ == '__main__':
	get_degree_days()
